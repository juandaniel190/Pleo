"""
Amezquita_Daniel_RevOps — Part 4 Python deliverable
====================================================

Purpose
-------
1. Read the raw Pleo Excel workbook
2. Validate each tab (row counts, PK uniqueness, FK referential integrity,
   null counts on critical fields, basic schema sanity)
3. Write one CSV per source table into the dbt_pleo/seeds/ folder so that
   `dbt seed` can load them into DuckDB

This script is the single source of ingestion. The Excel file is treated as the
raw landing zone; the CSVs that drop into `seeds/` are the immutable, versioned
inputs to the modelling layer.

The script intentionally does NOT clean the data. Cleaning is the job of the
intermediate layer in dbt. Here we only:
  - locate the header row in each tab (Excel has a title + description above
    the header — this is intentional in the source workbook)
  - keep raw column names exactly as-is so staging models can map them
  - emit a validation report so issues are visible before modelling

Run
---
    python deliverables/04_python_script.py

Outputs
-------
    dbt_pleo/seeds/raw_<table>.csv  (×7)
    deliverables/validation_report.json
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path("/sessions/amazing-sweet-darwin/mnt/uploads/Pleo RevOps Challenge Databases.xlsx")
SEEDS_DIR = ROOT / "dbt_pleo" / "seeds"
REPORT_PATH = ROOT / "deliverables" / "validation_report.json"

# Map of Excel tab → seed file stem. Order matters for FK validation.
SHEETS = {
    "Accounts": "raw_accounts",
    "Rep Roster": "raw_rep_roster",
    "Opportunities": "raw_opportunities",
    "Pipeline Snapshots": "raw_pipeline_snapshots",
    "Funnel Events": "raw_funnel_events",
    "Sales Targets": "raw_sales_targets",
    "Activity Data": "raw_activity_data",
}

# Primary keys for uniqueness checks (None means the table has a composite/no PK)
PRIMARY_KEYS = {
    "raw_accounts": "account_id",
    "raw_rep_roster": "rep_id",
    "raw_opportunities": "opportunity_id",
    "raw_funnel_events": "event_id",
    "raw_sales_targets": "target_id",
    "raw_activity_data": "activity_id",
    "raw_pipeline_snapshots": None,  # composite: snapshot_month + opportunity_id
}

# Foreign keys to check (child_table, child_col) → (parent_table, parent_col)
FOREIGN_KEYS = [
    ("raw_opportunities", "account_id", "raw_accounts", "account_id"),
    ("raw_opportunities", "owner_id", "raw_rep_roster", "rep_id"),
    ("raw_pipeline_snapshots", "opportunity_id", "raw_opportunities", "opportunity_id"),
    ("raw_pipeline_snapshots", "owner_id", "raw_rep_roster", "rep_id"),
    ("raw_funnel_events", "account_id", "raw_accounts", "account_id"),
    ("raw_funnel_events", "rep_id", "raw_rep_roster", "rep_id"),
    ("raw_sales_targets", "rep_id", "raw_rep_roster", "rep_id"),
    ("raw_activity_data", "rep_id", "raw_rep_roster", "rep_id"),
    ("raw_activity_data", "account_id", "raw_accounts", "account_id"),
    ("raw_activity_data", "opportunity_id", "raw_opportunities", "opportunity_id"),
]


def find_header_row(ws) -> tuple[int, list[str]]:
    """Header row = first row that looks like snake_case field names.

    The Excel has two intentional metadata rows above the header (a title and a
    description). The header is the first row where most cells are non-null
    strings and contain an underscore.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_null = [c for c in row if c is not None]
        if len(non_null) < 4:
            continue
        if all(isinstance(c, str) for c in non_null) and any("_" in c for c in non_null):
            return i, non_null
    raise ValueError(f"Could not find header row in {ws.title}")


def extract_sheet(wb, sheet_name: str) -> pd.DataFrame:
    """Read one sheet into a DataFrame, locating the header row dynamically."""
    ws = wb[sheet_name]
    header_idx, headers = find_header_row(ws)
    n_cols = len(headers)
    rows = []
    for row in ws.iter_rows(min_row=header_idx + 1, max_col=n_cols, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=headers)
    return df


def validate(dfs: dict[str, pd.DataFrame]) -> dict:
    """Run the basic validation suite and return a structured report."""
    report: dict = {"row_counts": {}, "pk_uniqueness": {}, "null_counts": {}, "fk_violations": {}}

    for name, df in dfs.items():
        report["row_counts"][name] = len(df)
        # null counts (only show columns with at least one null)
        nulls = df.isna().sum()
        report["null_counts"][name] = {c: int(n) for c, n in nulls.items() if n > 0}
        # PK uniqueness
        pk = PRIMARY_KEYS.get(name)
        if pk:
            dups = df[df.duplicated(subset=[pk], keep=False)]
            report["pk_uniqueness"][name] = {
                "pk": pk,
                "duplicate_rows": int(len(dups)),
                "duplicate_pk_values": dups[pk].astype(str).unique().tolist() if len(dups) else [],
            }

    # FK referential integrity
    for child_tbl, child_col, parent_tbl, parent_col in FOREIGN_KEYS:
        child = dfs[child_tbl]
        parent = dfs[parent_tbl]
        # Allow nulls in child (e.g. activity.opportunity_id is legitimately null for SDR pre-opp work)
        child_vals = child[child_col].dropna().astype(str)
        parent_vals = set(parent[parent_col].dropna().astype(str))
        orphans = child_vals[~child_vals.isin(parent_vals)]
        key = f"{child_tbl}.{child_col} -> {parent_tbl}.{parent_col}"
        report["fk_violations"][key] = {
            "orphan_count": int(orphans.shape[0]),
            "orphan_values_sample": orphans.unique()[:10].tolist(),
        }

    return report


def main() -> None:
    print(f"Reading {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    SEEDS_DIR.mkdir(parents=True, exist_ok=True)

    dfs: dict[str, pd.DataFrame] = {}
    for sheet_name, seed_stem in SHEETS.items():
        df = extract_sheet(wb, sheet_name)

        # We deliberately keep mixed date formats and `%`-containing percentages
        # as-is in the CSVs. dbt staging models handle the cleaning via the
        # `parse_date` and `parse_percent` macros so the transformation logic
        # lives in version-controlled SQL alongside the rest of the model layer.

        dfs[seed_stem] = df
        out = SEEDS_DIR / f"{seed_stem}.csv"
        df.to_csv(out, index=False)
        print(f"  wrote {out.name:35s}  {len(df):>4d} rows  {len(df.columns):>2d} cols")

    report = validate(dfs)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(json.dumps(report, indent=2, default=str))
    print(f"\nValidation report → {REPORT_PATH}")

    # Print headline issues to stdout
    print("\n--- VALIDATION SUMMARY ---")
    for tbl, info in report["pk_uniqueness"].items():
        if info["duplicate_rows"]:
            print(f"  ⚠ {tbl}: {info['duplicate_rows']} duplicate rows on {info['pk']}: {info['duplicate_pk_values']}")
    for key, info in report["fk_violations"].items():
        if info["orphan_count"]:
            print(f"  ⚠ FK orphans {key}: {info['orphan_count']} ({info['orphan_values_sample']})")
    print("  (any other warnings above are acceptable nulls; see validation_report.json)")


if __name__ == "__main__":
    main()
